# coding: utf-8
import cmd
import json
import logging
import shutil
import datetime
import os
from enum import Enum, unique

import openpyxl
import pytz
import urllib3
from huaweicloudsdkcore.auth.credentials import BasicCredentials, GlobalCredentials
from huaweicloudsdkcore.http.http_config import HttpConfig
from huaweicloudsdkcore.exceptions import exceptions
from huaweicloudsdkiam.v3 import IamClient, KeystoneListUsersRequest
from huaweicloudsdkswr.v2 import *

logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

log_path = './logs'
if not os.path.exists(log_path):
    os.makedirs(log_path)
file_handler = logging.FileHandler(log_path + os.sep + 'swr-batch.log')
file_handler.setLevel(logging.INFO)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter(
    '%(asctime)s][%(name)-5s][%(levelname)-5s][%(funcName)s] %(message)s (%(filename)s:%(lineno)d)')
file_handler.setFormatter(formatter)
console_handler.setFormatter(formatter)

logger.addHandler(file_handler)
logger.addHandler(console_handler)

VERSION = "v1.02"


@unique
class AUTH_OPTION(Enum):
    SELF = 1
    OTHERS = 2


@unique
class IMG_RETAINS_OPTION(Enum):
    BY_DAY = 1
    BY_NUM = 2


RESULT_FAIL = "fail"
RESULT_SUCCESS = "success"


class SWR_BATCH(object):
    def __init__(self, ak, sk, swr_endpoint,iam_endpoint,domain_id):
        super().__init__()
        self.ak = ak
        self.sk = sk
        self.swr_endpoint = swr_endpoint
        self.iam_endpoint = iam_endpoint
        self.domain_id = domain_id
        self.swr_client = self.get_swr_client()
        self.iam_client = self.get_iam_client()

    def get_swr_client(self):
        client = None
        try:
            credentials = BasicCredentials(self.ak, self.sk)
            client = SwrClient.new_builder() \
                .with_credentials(credentials) \
                .with_http_config(HttpConfig(ignore_ssl_verification=True)) \
                .with_endpoint(self.swr_endpoint) \
                .build()
        except Exception as e:
            logger.error("获取swr客户端异常，ak :{}".format(ak))
            logger.error(e)
        return client

    def get_iam_client(self):
        client = None
        try:
            credentials = GlobalCredentials(self.ak, self.sk) \
                .with_domain_id(self.domain_id)
            client = IamClient.new_builder() \
                .with_credentials(credentials) \
                .with_http_config(HttpConfig(ignore_ssl_verification=True)) \
                .with_endpoint(self.iam_endpoint) \
                .build()
        except Exception as e:
            logger.error("获取iam客户端异常，ak :{}".format(ak))
            logger.error(e)
        return client

    #需要iam管理员角色才能查询
    def get_iam_user_list(self):
        ret =[]
        try:
            request = KeystoneListUsersRequest()
            request.enabled = True
            response = self.iam_client.keystone_list_users(request)
            if response and response.users:
                ret=response.users
        except exceptions.ClientRequestException as e:
            logger.error("get_iam_users error ! status_code {}  ,request_id {}  ,error_code {},  error_msg: {}" \
                         .format(e.status_code,e.request_id,e.error_code,e.error_msg))
        return ret

    def get_repository(self, namespace):
        ret = []
        try:
            request = ListReposDetailsRequest(namespace=namespace)
            response = self.swr_client.list_repos_details(request)
            if response and response.body:
                ret = response.body
        except exceptions.ClientRequestException as e:
            logger.error("get_repository error ! , namespace is  {} ,status_code {}  ,request_id {}  ,error_code {},  error_msg: {}" \
                    .format(namespace, e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret

    #获取全部用户 self_auth + others_auths
    def get_img_auth_list(self,namespace_name,repository_name):
        ret = []
        try:
            request = ShowUserRepositoryAuthRequest()
            request.namespace = namespace_name
            request.repository = self.fix_repo_name(repository_name)
            response = self.swr_client.show_user_repository_auth(request)
            if response and response.others_auths:
                ret = response.others_auths
            ret.append(response.self_auth)
        except exceptions.ClientRequestException as e:
            logger.error("get_repository_user_auth_list error,namespace:{},repository:{}" \
                         .format(namespace_name, repository_name))
            logger.error(" status_code {}  ,request_id {}  ,error_code {},  error_msg: {}" \
                .format(e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret

    def update_img_auth(self, namespace_name, repository_name, user_id,user_name,auth_value):
        ret = -1
        try:
            if isinstance(auth_value, str):
                auth_value = int(auth_value)
            request = UpdateUserRepositoryAuthRequest()
            request.namespace = namespace_name
            request.repository = self.fix_repo_name(repository_name)
            body = [
                UserAuth(
                    user_id=user_id,
                    user_name=user_name,
                    auth=auth_value
                )
            ]
            request.body = body
            self.swr_client.update_user_repository_auth(request)
            ret = 1
            logger.info( "update_img_auth OK ! org: {}  repository {} user_id {}  user_name:  {}  auth_value {} " \
                         .format(namespace_name, repository_name, user_id,user_name, auth_value))
        except exceptions.ClientRequestException as e:
            logger.error( "update_img_auth ERROR !, org: {}  repository {} user_id {}  user_name:  {}  auth_value {} " \
                         .format(namespace_name, repository_name, user_id,user_name, auth_value))
            logger.error("status_code {} ,request_id {},error_code {},error_msg {}" \
                         .format(e.status_code,e.request_id,e.error_code,e.error_msg))
        return ret


    def create_img_auth(self,namespace_name,repository_name,user_id,user_name,auth_value):
        ret = -1
        try:
            if isinstance(auth_value, str):
                auth_value = int(auth_value)
            request = CreateUserRepositoryAuthRequest()
            request.namespace = namespace_name
            request.repository = self.fix_repo_name(repository_name)
            body = [
                UserAuth(
                    user_id=user_id,
                    user_name=user_name,
                    auth=auth_value
                )
            ]
            request.body = body
            self.swr_client.create_user_repository_auth(request)
            ret =1
            logger.info("create_img_auth OK!,namespace {} ,repository {}  ,user_id {},user_name {},auth_value {}" \
                        .format(namespace_name,repository_name,user_id, user_name,auth_value))
        except exceptions.ClientRequestException as e:
            logger.error("create_img_auth error!,namespace {}, repository {}  ,user_id {},user_name {},auth_value {}" \
                         .format(namespace_name,repository_name,user_id,user_name,auth_value))
            logger.error("status_code {},request_id {},error_code {},error_msg{}".format(e.status_code,e.request_id,e.error_code,e.error_msg))
        return ret


    def deal_with_img_auth(self,namespace_name, repository_name, user_id,user_name,auth_value):
        ret = -1
        namespace_list = [ns.name.strip() for ns in self.get_namespace()]
        auth_dict={}
        iam_all_user_list=self.get_iam_user_list()

        for auth in self.get_img_auth_list(namespace_name, repository_name):
            auth_dict[auth.user_name] = auth.user_id

        # 判断是否已经在镜像授权里面
        if user_name in auth_dict.keys():
            # 在镜像授权里面 修改授权
            ret = self.update_img_auth(namespace_name, repository_name, auth_dict[user_name], user_name, auth_value)
        else:
            # 不在镜像授权里面 新增授权
            ret = self.create_img_auth(namespace_name, repository_name, user_id, user_name, auth_value)
        return ret

    def batch_set_img_auth(self,excel_path):
        now_time_str = COMMON.now_time_str()
        namespace_list = [ns.name.strip() for ns in self.get_namespace()]
        iam_all_user_list = self.get_iam_user_list()
        iam_user_dict={}
        for iam_user in iam_all_user_list:
            iam_user_dict[iam_user.name]=iam_user.id
        excel_rows = self.read_excel(excel_path)
        row_count = 1
        for row in excel_rows: #逐行处理excel表格
            logger.info("--->process a row  ")
            row_count += 1
            self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_FAIL)
            try:
                namespace_name = self.convert_trim_str(row[0])
                # 判断namespace是否存在
                if namespace_name not in namespace_list:
                    logger.error("organization {} not found.skip !".format(namespace_name))
                    continue
                # 判断repository 是否存在
                repository_name = self.convert_trim_str(row[1])
                repository_list = [repo.name for repo in self.get_repository(namespace_name)]
                if repository_name not in repository_list:
                    logger.error("organization {} -> repository {} not found.skip!".format(namespace_name, repository_name))
                    continue
                # 判断人是否iam的用户
                user_name = self.convert_trim_str(row[2])
                if  user_name not in iam_user_dict.keys():
                    continue
                #判断权限是否正确
                auth_value = row[3]
                if isinstance(auth_value, str):
                    auth_value = int(auth_value)
                if auth_value not in [1,3,7]:
                    continue
                ret=self.deal_with_img_auth(namespace_name, repository_name, iam_user_dict[user_name], user_name, auth_value)
                if ret and ret > 0:
                    self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_SUCCESS)
            except Exception as e:
                logger.error("process row data error!  row-data:{}".format(row))
                logger.error(e)
            logger.info("<---process  a row  done")

    def deal_with_org_auth(self, namespace_name, user_id, user_name, auth_value):
        ret = -1
        # 查询当前组织权限是否有这个人
        exists = False
        for org_user_auth in self.get_namespace_user_auth_list(namespace_name):
            if org_user_auth.user_name == user_name:
                exists = True
                break
        # 人在组织权限用户中，修改他
        if exists:
            ret = self.update_org_auth(namespace_name, user_id, user_name, auth_value)
        else:
            # 人在iam系统中，但是没有在当前权限列表中，创建权限
            ret = self.create_org_auth(namespace_name, user_id, user_name, auth_value)
        return ret

    def deal_with_repo_ageing(self, namespace_name, repo_name, template, num, label_value, regex_pattern):
        ret = -1
        retentions = self.get_repository_retentions(namespace_name, repo_name)
        if retentions:
            # 查看template类型是否一致，不一致跳过，一致的化更新保留策略
            if retentions and retentions[0] and retentions[0].id and retentions[0].rules[0] and retentions[0] \
                .rules[0].template:
                id = retentions[0].id
                template_from_sys = retentions[0].rules[0].template
                if template_from_sys == template:
                    ret = self.update_repository_retentions(namespace_name, repo_name, id, template_from_sys, num,
                                                            label_value, regex_pattern)
                else:
                    logger.error("namespace {}  repository {} ,retentions exist , update error not allow modify template(such as :tag<-->date) ,skip" \
                        .format(namespace_name, repo_name))
                    return ret
            else:
                logger.error("namespace {}  repository {} ,retentions exist ,but error ,skip" \
                             .format(namespace_name, repo_name))
                return ret
        else:
            logger.info("namespace {}  repository {} , retentions not exist ,will create " \
                        .format(namespace_name, repo_name))
            ret = self.create_repository_retentions(namespace_name, repo_name, template, num, label_value,
                                                    regex_pattern)
        return ret

    def get_int_value(self, origin_value):
        ret=origin_value
        if isinstance(origin_value, str):
            ret=int(origin_value)
        return ret

    def create_org_auth(self,namespace_name,user_id,user_name,auth_value):
        ret =-1
        try:
            request = CreateNamespaceAuthRequest()
            request.namespace = namespace_name
            listBodybody = [
                UserAuth(
                    user_id=user_id,
                    user_name=user_name,
                    auth=self.get_int_value(auth_value)
                )
            ]
            request.body = listBodybody
            self.swr_client.create_namespace_auth(request)
            ret =1
            logger.info("create org auth success ,org :{} ,user_name: {}, auth_value :{}".format(namespace_name,user_name,auth_value))
        except exceptions.ClientRequestException as e:
            logger.error("create org auth error ,org :{} ,user_id : {},user_name: {}, auth_value :{}" \
                    .format(namespace_name,user_id, user_name,auth_value))
            logger.error("status_code {}  ,request_id {}  ,error_code {},  error_msg: {}" \
                .format(e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret

    def update_org_auth(self,namespace_name,user_id,user_name,auth_value):
        ret =-1
        try:
            request = UpdateNamespaceAuthRequest()
            request.namespace = namespace_name
            listBodybody = [
                UserAuth(
                    user_id=user_id,
                    user_name=user_name,
                    auth=self.get_int_value(auth_value)
                )
            ]
            request.body = listBodybody
            self.swr_client.update_namespace_auth(request)
            ret =1
            logger.info("update_org_auth success ,namespace_name {} user_id {} user_name {} auth_value :{}".format(namespace_name,user_id,user_name,auth_value))
        except exceptions.ClientRequestException as e:
            logger.error("update_org_auth error ,namespace_name {} user_id {} user_name {} auth_value :{}".format(namespace_name,user_id,user_name,auth_value))
            logger.error(" status_code {}  ,request_id {}  ,error_code {},  error_msg: {}".format(
                e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret

    def create_repository_auth(self):
        pass

    def update_repository_auth(self):
        pass

    def read_excel(self, excel_path):
        ret = []
        work_book = openpyxl.load_workbook(excel_path)
        sheet_names = work_book.sheetnames
        sheet = work_book[sheet_names[0]]

        head_row = 1
        for row in sheet.iter_rows(min_row=head_row + 1, values_only=True):
            ret.append(row)
        return ret

    def fix_repo_name(self, repository):
        if repository.count("/") > 0:
            repository = repository.replace("/", "$")
        return repository

    def get_repository_tag(self, namespace, repository):
        tag_list = []
        try:
            request = ListRepositoryTagsRequest()
            request.namespace = namespace
            request.repository = self.fix_repo_name(repository)
            response = self.swr_client.list_repository_tags(request)
            if response and response.body:
                tag_list = response.body
        except exceptions.ClientRequestException as e:
            logger.error("get_repository_tag error,namespace:{},repository:{}".format(namespace, repository))
            logger.error(" status_code {}  ,request_id {}  ,error_code {},  error_msg: {}".format(
                e.status_code, e.request_id, e.error_code, e.error_msg))
        return tag_list

    def get_namespace_user_auth_list(self, namespace):
        auth_users = self.get_namespace_auth(namespace, AUTH_OPTION.OTHERS)
        auth_users_auth_list = []
        for auth_user in auth_users:
            auth_users_auth_list.append(auth_user)
        return auth_users_auth_list

    def get_repository_user_auth_list(self, namespace, repository):
        ret = []
        try:
            request = ShowUserRepositoryAuthRequest()
            request.namespace = namespace
            request.repository = repository
            response = self.swr_client.show_user_repository_auth(request)
            if response:
                ret = response.others_auths
        except exceptions.ClientRequestException as e:
            logger.error("get_repository_user_auth_list error,namespace:{},repository:{}".format(namespace, repository))
            logger.error(" status_code {}  ,request_id {}  ,error_code {},  error_msg: {}".format(
                e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret

    def set_image_user_auth(self, excel_path):
        now_time_str = COMMON.now_time_str()
        namespace_list = [ns.name.strip() for ns in self.get_namespace()]
        excel_rows = self.read_excel(excel_path)
        user_list_cache = {}
        row_count = 1
        for row in excel_rows:
            row_count += 1
            self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_FAIL)
            try:
                namespace = self.convert_trim_str(row[0])
                # 判断namespace是否存在
                if namespace not in namespace_list:
                    logger.error("organization {} not found. skip!".format(namespace))
                    continue

                repository = self.convert_trim_str(row[1])
                user = self.convert_trim_str(row[2])
                auth = row[3]
                if isinstance(auth, str):
                    auth = int(auth.strip())

                # 判断repository 是否存在
                repository_list = [repo.name for repo in self.get_repository(namespace)]
                if repository not in repository_list:
                    logger.error("organization {} -> repository {} not found. skip!".format(namespace, repository))
                    continue

                # 获取此namespace下的user_auth
                target_auth_list = []
                if namespace in user_list_cache:
                    user_auth_list = user_list_cache[namespace]
                else:
                    user_auth_list = self.get_repository_user_auth_list(namespace, repository)
                    user_list_cache[namespace] = user_auth_list

                # 判断用户是否存在
                user_auth_name_list = [u.user_name.strip() for u in user_auth_list]
                if user not in user_auth_name_list:
                    logger.error(
                        "organization {} ,repository {}  -> user {} not found or can not modify for the user.".format(
                            namespace, repository, user))
                    continue

                # excel中该行的 namespace repo user 都存在的前提下修改权限，人数不多不会出现性能瓶颈
                for ua in user_auth_list:
                    if user == ua.user_name.strip():
                        ua.auth = auth
                        target_auth_list.append(ua)
                        break
                if target_auth_list:
                    ret = self.modify_repository_user_auth(namespace, repository, target_auth_list)
                    if ret and ret > 0:
                        self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_SUCCESS)

                else:
                    logger.error("image_user_auth error ")
            except Exception as e:
                logger.error(e)

    def set_org_user_auth(self, excel_path):
        now_time_str = COMMON.now_time_str()
        namespace_list = [ns.name.strip() for ns in self.get_namespace()]
        iam_user_list=self.get_iam_user_list()
        excel_rows = self.read_excel(excel_path)
        row_count = 1
        for row in excel_rows:
            logger.info("--->process a row ")
            row_count += 1
            self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_FAIL)
            try:
                namespace = self.convert_trim_str(row[0])
                # 判断namespace是否存在
                if namespace not in namespace_list:
                    logger.error("organization {} not found. skip!".format(namespace))
                    continue
                # 判断username是否是合法的iam用户
                user_name = self.convert_trim_str(row[1])
                target_user = None
                for iam_user in iam_user_list:
                    if iam_user.name == user_name:
                        target_user = iam_user
                        break
                if not target_user:
                    logger.error("username : {}  error, not in iam system ,skip ! ".format(user_name))
                    continue
                #获取权限值
                auth_value = row[2]
                if isinstance(auth_value, str):
                    auth_value = int(auth_value.strip())
                if not auth_value in [1,3,7] :
                    logger.error("namespace {} username  {}  auth {} , but auth_value is error  :   auth only 1 or 3 or 7 ,skip ! ".format(namespace,user_name, auth_value))
                    continue

                ret=self.deal_with_org_auth(namespace,target_user.id,user_name,auth_value)
                if ret and ret > 0:
                    self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_SUCCESS)
                else:
                    logger.error("log_result_to_excel error ,ret =< 0 ")
            except Exception as e:
                logger.error(e)
            logger.info("<---process  a row done")

    def convert_trim_str(self, value):
        ret = value
        if not isinstance(value, str):
            try:
                ret = str(value)
            except Exception as e:
                logger.info("excel cell  convert to string  error ,cell value is {}".format(value))
        return ret.strip()

    def bat_ageing_img(self, excel_path):
        now_time_str = COMMON.now_time_str()
        namespace_list = [ns.name.strip() for ns in self.get_namespace()]
        excel_rows = self.read_excel(excel_path)
        row_count = 1
        for row in excel_rows:
            row_count += 1
            self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_FAIL)
            namespace=""
            repository=""
            try:
                namespace = self.convert_trim_str(row[0])
                # 判断namespace是否存在
                if namespace not in namespace_list:
                    logger.error("organization {} not found.".format(namespace))
                    continue
                repository = self.convert_trim_str(row[1])
                template = self.convert_trim_str(row[2])
                num = self.convert_trim_str(row[3])
                filter_label = self.convert_trim_str(row[4])
                filter_regex = self.convert_trim_str(row[5])
                # 判断repository 是否存在
                repository_list = [repo.name for repo in self.get_repository(namespace)]
                if repository not in repository_list:
                    logger.error("organization {}  repository {} not found.".format(namespace, repository))
                    continue
                #判断template是否合法
                if template.lower() not in ["date_rule", "tag_rule"]:
                    logger.error("template is {} error. only : date_rule or tag_rule".format(template))
                    continue
                result = self.deal_with_repo_ageing(namespace, repository, template.lower(), num, filter_label, filter_regex)
                if result and result > 0:
                    self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_SUCCESS)
            except Exception as e:
                logger.error("batch set ageing policy error,namespace:{},repository:{}".format(namespace, repository))
                logger.error(e)

    def batch_delete_tag(self, excel_path):
        now_time_str = COMMON.now_time_str()
        namespace_list = [ns.name.strip() for ns in self.get_namespace()]
        excel_rows = self.read_excel(excel_path)
        row_count = 1
        for row in excel_rows:
            logger.info("--->process a row  ")
            row_count += 1
            self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_FAIL)
            try:
                namespace = self.convert_trim_str(row[0])
                # 判断namespace是否存在
                if namespace not in namespace_list:
                    logger.error("organization {} not found.skip !".format(namespace))
                    continue
                namespace = self.convert_trim_str(row[0])
                repository = self.convert_trim_str(row[1])
                tag = self.convert_trim_str(row[2])
                is_delete = self.convert_trim_str(row[3])
                # 判断repository 是否存在
                repository_list = [repo.name for repo in self.get_repository(namespace)]
                if repository not in repository_list:
                    logger.error("organization {} -> repository {} not found.skip!".format(namespace, repository))
                    continue
                # 判断是否存在tag
                tag_list = self.get_repository_tag(namespace, repository)
                all_tag = [tag.tag for tag in tag_list]
                if tag not in all_tag:
                    logger.error(
                        "organization {} , repository {} -> tag {} not found.skip!".format(namespace, repository, tag))
                    continue
                if is_delete.lower() == "y":
                    ret = self.delete_image_tag(namespace, repository, tag)
                    if ret and ret > 0:
                        self.log_result_to_excel(excel_path, row_count, now_time_str, RESULT_SUCCESS)
                else:
                    logger.info( "organization {}  repository {}  tag {}  , delete_flag is not Y or y. skip!".format(namespace, repository, tag))
            except Exception as e:
                logger.error(e)
            logger.info("<---process  a row  done")

    def get_namespace(self):
        ret = []
        try:
            request = ListNamespacesRequest()
            response = self.swr_client.list_namespaces(request)
            if response and response.namespaces:
                ret = response.namespaces
        except exceptions.ClientRequestException as e:
            logger.error(" status_code {}  ,request_id {}  ,error_code {},  error_msg: {}".format(
                e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret

    def get_namespace_auth(self, namespace, auth_option):
        ret = []
        try:
            request = ShowNamespaceAuthRequest()
            request.namespace = namespace
            response = self.swr_client.show_namespace_auth(request)
            if auth_option == AUTH_OPTION.OTHERS and response and response.others_auths:
                ret = response.others_auths
            elif auth_option == AUTH_OPTION.SELF and response and response.self_auth:
                ret = response.self_auth
        except exceptions.ClientRequestException as e:
            logger.error(" status_code {}  ,request_id {}  ,error_code {},  error_msg: {}".format(
                e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret

    def modify_namespace_user_auth(self, namespace, user_auth):
        ret = -1
        try:
            request = UpdateNamespaceAuthRequest()
            request.namespace = namespace
            request.body = user_auth
            self.swr_client.update_namespace_auth(request)
            ret = 1
            logger.info(
                "org: " + namespace + " user: " + user_auth[0].user_name + " auth: " + str(
                    user_auth[0].auth) + " OK!")
        except exceptions.ClientRequestException as e:
            logger.error("modify_namespace_user_auth err!")
            logger.error(
                "org: " + namespace + " user: " + user_auth[0].user_name + " auth: " + str(
                    user_auth[0].auth) + " ERROR!")
            logger.error(" status_code {}  ,request_id {}  ,error_code {},  error_msg: {}".format(
                e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret

    def modify_repository_user_auth(self, namespace, repository, user_auth):
        ret = -1
        try:
            request = UpdateUserRepositoryAuthRequest()
            request.namespace = namespace
            request.repository = repository
            request.body = user_auth
            self.swr_client.update_user_repository_auth(request)
            ret = 1
            logger.info(
                "org: " + namespace + " repository: " + repository + " user: " + user_auth[
                    0].user_name + " auth: " + str(user_auth[0].auth) + " OK!")
        except exceptions.ClientRequestException as e:
            logger.error("org: " + namespace + " repository: " + repository + " user: " + user_auth[
                0].user_name + " auth: " + str(user_auth[0].auth) + " ERROR!")
            logger.error(" status_code {}  ,request_id {}  ,error_code {},  error_msg: {}".format(
                e.status_code, e.request_id, e.error_code, e.error_msg))
        return ret


    def log_result_to_excel(self, excel_path, current_row_count, now_time_str, result_flag):
        try:
            # copy一份excel,添加一列result
            excel_name = os.path.basename(excel_path)
            excel_dir = os.path.dirname(excel_path)
            log_excel_dir = excel_dir + os.sep + "logs"

            if not os.path.exists(log_excel_dir):
                os.makedirs(log_excel_dir)
            log_excel_file = log_excel_dir + os.sep + "log." + now_time_str + "." + excel_name
            logger.info("log_excel_file: {}".format(log_excel_file))
            if not os.path.exists(log_excel_file):
                shutil.copy(excel_path, log_excel_file)
                workbook = openpyxl.load_workbook(log_excel_file)
                sheet = workbook.active
                # 写入表头
                sheet.cell(row=1, column=sheet.max_column + 1, value="result")
                workbook.save(log_excel_file)
                workbook.close()

            # 如果新创建的日志文件，给日志excel添加一列结果列,都置为失败
            workbook = openpyxl.load_workbook(log_excel_file)
            sheet = workbook.active
            sheet.cell(row=current_row_count, column=sheet.max_column, value=result_flag)
            workbook.save(log_excel_file)
            workbook.close()
        except Exception as e:
            logger.error("log result to excel error ,excel_path: {}".format(excel_path))
            logger.error(e)

    def delete_image_tag(self, namespace, repository, tag):
        ret = -1
        try:
            request = DeleteRepoTagRequest()
            request.namespace = namespace
            request.repository = repository
            request.tag = tag
            self.swr_client.delete_repo_tag(request)
            ret = 1
            logger.info(
                "** namespace {} repository {} tag {}  delete_image_tag  OK! **".format(namespace, repository, tag))
        except Exception as e:
            logger.error(
                "** namespace {} repository {} tag {}  delete_image_tag  ERROR! **".format(namespace, repository, tag))
            logger.error(e)
        return ret

    def get_repository_retentions(self, namespace_name, repository_name):
        ret = []
        try:
            request = ListRetentionsRequest()
            request.namespace = namespace_name
            request.repository = self.fix_repo_name(repository_name)
            response = self.swr_client.list_retentions(request)
            if response and response.body:
                ret = response.body
        except exceptions.ClientRequestException as e:
            logger.error("get_repository_retention error,namespace {},repository {} ".format(namespace_name, repository_name))
            logger.error(" status_code: {} ,request_id {} ,error_code {} ,error_msg {}".format(e.status_code,e.request_id,e.error_code,e.error_msg)  )
        return ret
    # template :date_rule,tag_rule
    # num： date_rule num天，tag_rule num 个
    # kind:label、regexp
    # pattern:kind是label时,设置为镜像版本,kind是regexp时,设置为正则表达式
    def create_repository_retentions(self, namespace_name, repository_name, template, num, label_value,regex_pattern):
        ret =-1
        try:
            request = CreateRetentionRequest()
            request.namespace = namespace_name
            request.repository = self.fix_repo_name(repository_name)
            tag_filter = []
            if label_value:
                labels=label_value.split(",")
                for label in labels:
                    tag_filter.append(TagSelector(
                        kind="label",
                        pattern=label
                    ))
            if regex_pattern:
                tag_filter.append(TagSelector(
                    kind="regexp",
                    pattern=regex_pattern
                ))


            listRulesbody=[]
            if template.lower()=="date_rule":
                listRulesbody = [
                    Rule(
                        template=template,
                        params={"days": str(num)},
                        tag_selectors=tag_filter
                    )
                ]
            elif template.lower()=="tag_rule":
                listRulesbody = [
                    Rule(
                        template=template,
                        params={"num": str(num)},
                        tag_selectors=tag_filter
                    )
                ]

            request.body = CreateRetentionRequestBody(
                rules=listRulesbody,
                algorithm="or"
            )
            self.swr_client.create_retention(request)
            ret =1
            logger.error("create_repository_retentions OK!,namespace {},repository {} ".format(namespace_name,
                                                                                                 repository_name))
        except exceptions.ClientRequestException as e:
            logger.error("create_repository_retentions error,namespace {},repository {} ".format(namespace_name,
                                                                                                 repository_name))
            logger.error(
                " status_code {} ,request_id {} , error_code{} ,error_msg {}".format(e.status_code, e.request_id,
                                                                                   e.error_code, e.error_msg))
        return ret

    # retention_id ,template 是查询后赋值的
    # template不允许修改
    def update_repository_retentions(self, namespace_name, repository_name,retention_id,template, num, label_value,regex_pattern):
        ret =-1
        try:
            request = UpdateRetentionRequest()
            request.namespace = namespace_name
            request.repository = self.fix_repo_name(repository_name)
            if isinstance(retention_id,str):
                retention_id=int(retention_id)
            request.retention_id = retention_id
            tag_filter=[]
            if label_value:
                labels = label_value.split(",")
                for label in labels:
                    tag_filter.append(TagSelector(
                        kind="label",
                        pattern=label
                    ))
            if regex_pattern:
                tag_filter.append(TagSelector(
                    kind="regexp",
                    pattern=regex_pattern
                ))

            listRulesbody = []
            if template.lower() == "date_rule":
                listRulesbody = [
                    Rule(
                        template=template,
                        params={"days": str(num)},
                        tag_selectors=tag_filter
                    )
                ]
            elif template.lower() == "tag_rule":
                listRulesbody = [
                    Rule(
                        template=template,
                        params={"num": str(num)},
                        tag_selectors=tag_filter
                    )
                ]
            request.body = UpdateRetentionRequestBody(
                rules=listRulesbody,
                algorithm="or"
            )
            self.swr_client.update_retention(request)
            ret = 1
            logger.info("update_repository_retentions OK!,namespace {},repository {} ,template {},num {},label_value {},regex_pattern {}" \
                .format(namespace_name, repository_name,template,num,label_value,regex_pattern))
        except exceptions.ClientRequestException as e:
            logger.error("update_repository_retentions ERROR!,namespace {},repository {} ,template {},num {},label_value {},regex_pattern {}" \
                .format(namespace_name, repository_name,template,num,label_value,regex_pattern))
            logger.error(" status_code {} ,request_id {} , error_code{} ,error_msg {}" \
                         .format(e.status_code, e.request_id,e.error_code, e.error_msg))
        return ret

# 公共工具
class COMMON():
    @staticmethod
    def read_json( key):
        with open('auth-config.json', 'r', encoding='utf-8') as f:
            json_data = json.loads(f.read())
            return json_data[key]

    @staticmethod
    def now_time_str():
        return datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    @staticmethod
    def utc_to_local_time_str(utc_time_str):
        current_timezone = datetime.datetime.now(datetime.timezone.utc).astimezone().tzinfo
        try:
            utc_time = datetime.datetime.strptime(utc_time_str, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=pytz.UTC)
            current_time = utc_time.astimezone(current_timezone)
            return current_time.strftime("%Y-%m-%d %H:%M:%S")
        except Exception as e:
            logger.error("utc time convert local time error !,utc_time_str :{} ,current_timezone: {} ".format(utc_time_str,current_timezone))
            logger.error(e)
            return "-"


class SWR_BATCH_CLI(cmd.Cmd):
    prompt = 'swr-batch-cli> '

    def __init__(self, ak, sk, swr_endpoint,iam_endpoint,domain_id):
        super().__init__()
        self.swr_batch = SWR_BATCH(ak, sk, swr_endpoint,iam_endpoint,domain_id)

    def do_set_org_user(self, line):
        """batch set organization user auth by  excel . usage: set_org_user [excel-path]."""
        args = line.split()
        if len(args) != 1:
            print("usage: set_org_user [excel-path]")
        else:
            self.swr_batch.set_org_user_auth(args[0])

    def do_set_image_user(self, line):
        """batch set image user auth by  excel . usage: set_image_user [excel-path]."""
        args = line.split()
        if len(args) != 1:
            print("usage: set_image_user [excel-path]")
        else:
            self.swr_batch.set_image_user_auth(args[0])

    def do_set_image_ageing(self, line):
        """batch set image ageing by  excel . usage: set_image_ageing [excel-path]."""
        args = line.split()
        if len(args) != 1:
            print("usage: set_image_ageing [excel-path]")
        else:
            self.swr_batch.bat_ageing_img(args[0])

    def do_delete_tag(self, line):
        """batch delete tag by   excel . usage:  delete_tag [excel-path]."""
        args = line.split()
        if len(args) != 1:
            print("usage: delete_tag [excel-path]")
        else:
            self.swr_batch.batch_delete_tag(args[0])


    def do_quit(self, line):
        """Handle EOF (Ctrl-D or Ctrl-Z or quit) to exit."""
        return True

    def parseline(self, line):
        cmd, arg, line = super().parseline(line)
        if cmd:
            cmd = cmd.lower()
        return cmd, arg, line

    def onecmd(self, line):
        splits = line.split(None, 1)
        if not splits:
            return self.emptyline()
        command = splits[0].replace("-", "_")
        args = splits[1] if len(splits) > 1 else ''
        return super().onecmd(f"{command} {args}")

    def get_export_path(self):
        export_path = os.getcwd() + os.sep + "exports"
        if not os.path.exists(export_path):
            os.makedirs(export_path)
        return export_path
    def emptyline(self):
        pass

if __name__ == "__main__":
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    auth_config_path = "auth-config.json"
    ak = COMMON.read_json("HUAWEICLOUD_SDK_AK")
    sk = COMMON.read_json("HUAWEICLOUD_SDK_SK")
    swr_endpoint = COMMON.read_json("SWR_ENDPOINT")
    iam_endpoint = COMMON.read_json("IAM_ENDPOINT")
    domain_id = COMMON.read_json("DOMAIN_ID")

    # tip = "Welcome to swr-batch-cli(%s). Type 'help' or '?' to list commands. Type 'quit' to exit." % VERSION
    # app = SWR_BATCH_CLI(ak, sk, swr_endpoint, iam_endpoint, domain_id)
    # app.cmdloop(tip)

    #  below just for  test
    batch = SWR_BATCH(ak, sk, swr_endpoint, iam_endpoint, domain_id)

    # print(batch.get_iam_user_list())
    # batch.create_repository_retentions("test-libin","busybox","tag_rule",33,"v1,v3,v4,v6","regex-99")
    # batch.create_repository_retentions("test-libin","busybox","tag_rule",33,"v1,v3,v4,v6","regex-88")
    # batch.deal_with_repo_ageing("test-libin", "busybox",  "date_rule", 12, "v1,v3,v2", "regex-55")
    # batch.deal_with_repo_ageing("test-libin", "busybox",  "date_rule", 11, "v1", "regex-66")
    # batch.bat_ageing_img("/root/libin/image_retentions.xlsx")
    # print(batch.get_repository_retentions("test-libin", "busybox"))
    # batch.set_org_user_auth("/root/libin/org_user_auth.xlsx")

    # batch.get_repository_retentions("test-libin","busybox")
    # batch.create_org_auth("test-libin","5a65bfb9474c49a89710b19e0d84e6bd","qifei","1")
    # batch.deal_with_org_auth("test-libin","qifei","7")
    # batch.deal_with_org_auth("test-libin","xxx","3")
    # print(COMMON.now_time_str())
    # batch.get_iam_user_list()

    # batch.get_img_auth_list("test-libin","busybox")
    batch.batch_set_img_auth("/root/libin/img_user_auth.xlsx")

    print("done")
